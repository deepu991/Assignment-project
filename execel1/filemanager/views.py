import openpyxl
from django.shortcuts import render, redirect
from django.http import HttpResponse
from .forms import FileUploadForm
from .models import Order
from openpyxl import Workbook


def upload_file(request):
    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.cleaned_data['file']

            # Insert data from the Excel file into the database table
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                order_id, product_name, product_price, shipped = row

                Order.objects.create(order_id=order_id, product_name=product_name, product_price=product_price,
                                     shipped=shipped)

            return redirect('filemanager:upload')
    else:
        form = FileUploadForm()

    return render(request, 'upload.html', {'form': form})


def download_file(request):
    # Export all database details into an Excel file for download
    wb = Workbook()
    ws = wb.active

    # Add headers
    headers = ['Order ID', 'Product Name', 'Product Price', 'Shipped']
    ws.append(headers)

    # Add data
    orders = Order.objects.all()
    for order in orders:
        data = [order.order_id, order.product_name, order.product_price, order.shipped]
        ws.append(data)

    # Save the workbook to a temporary file
    filepath = 'temp.xlsx'
    wb.save(filepath)

    # Send the file as a response for download
    response = HttpResponse(open(filepath, 'rb'),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=orders.xlsx'

    return response
